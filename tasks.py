"""
Celery tasks for the video processing pipeline.

Converts the 9 pipeline stages into chained Celery tasks with:
- Progress reporting via self.update_state()
- Stage-level caching via Redis
- Cancellation and retry support
"""

import asyncio
import json
import os
import shutil
import subprocess
import tempfile
import urllib.request
from collections import namedtuple
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from tempfile import TemporaryDirectory
from typing import Optional
from urllib.parse import urljoin
from uuid import uuid4

import cv2
import m3u8
import pandas as pd
import skimage as ski
from billiard.exceptions import WorkerLostError
from celery import Task, chain
from celery.signals import task_failure
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

from cache import CacheContext, get_cached_result, set_cached_result
from celery_app import celery_app
# Import AI functions - they will run in sync context
from pipeline.ai import (
    clean_transcript,
    generate_captions,
    generate_spreadsheet_helper,
    generate_title,
    generate_vignette_questions,
)
from pipeline.helpers import Caption, Slide, fetch, parse_markdown_bold_to_rich_text


@task_failure.connect
def handle_task_failure(task_id, exception, args, kwargs, traceback, einfo, **kw):
    """Global handler for task failures including WorkerLostError."""
    if isinstance(exception, WorkerLostError):
        # Extract job_id from args
        job_id = None
        if args and isinstance(args[0], dict) and "job_id" in args[0]:
            job_id = args[0]["job_id"]
        elif args and len(args) > 0 and isinstance(args[0], int):
            job_id = args[0]

        if job_id:
            error_message = "Worker crashed (likely out of memory). Please try again with a shorter video."
            mark_job_failed(job_id, error_message)


class PipelineTask(Task):
    """Base task class with error handling for pipeline tasks."""

    def on_failure(self, exc, task_id, args, kwargs, einfo):
        """Handle task failure by updating job status in database."""
        # Extract job_id from args or kwargs
        job_id = None
        if args and isinstance(args[0], dict) and "job_id" in args[0]:
            job_id = args[0]["job_id"]
        elif args and len(args) > 0 and isinstance(args[0], int):
            job_id = args[0]

        if job_id:
            error_message = f"{exc.__class__.__name__}: {str(exc)}"
            mark_job_failed(job_id, error_message)


# Directory configuration
IN_DIR = os.path.join("data", "input")
OUT_DIR = os.path.join("data", "output")
FRAMES_DIR = os.path.join("data", "frames")

# Ensure directories exist
os.makedirs(IN_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)
os.makedirs(FRAMES_DIR, exist_ok=True)

# Named tuples for data structures
PanoptoInput = namedtuple("PanoptoInput", ["base", "cookie", "delivery_id"])


@dataclass
class TaskContext:
    """Context object that flows through the pipeline stages."""

    source_id: str
    input_type: str  # 'url', 'upload', 'panopto'
    input_data: dict  # Serialized input configuration
    use_ai: bool = True
    video_path: Optional[str] = None
    captions: Optional[list[dict]] = None
    slides: Optional[list[dict]] = None
    outputs: Optional[dict] = None

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> "TaskContext":
        return cls(**data)

    def get_video_path(self) -> str:
        """Get the path to the video file."""
        if self.video_path and os.path.exists(self.video_path):
            return self.video_path
        return os.path.join(IN_DIR, f"video_{self.source_id}.mp4")


def run_async(coro):
    """Run an async coroutine in a sync context."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


# Cached Flask app instance for Celery tasks
_flask_app = None


def _get_flask_app():
    """Get or create the Flask app instance for database operations."""
    global _flask_app
    if _flask_app is None:
        # Import here to avoid circular import at module load time
        from app import create_app

        _flask_app = create_app()
    return _flask_app


def update_job_progress(job_id: int, stage: str, progress: float, message: str) -> None:
    """Update job progress in the database."""
    from models import Job, JobStatus, db

    app = _get_flask_app()
    with app.app_context():
        job = db.session.get(Job, job_id)
        if job:
            job.current_stage = stage
            job.progress = progress
            if job.status == JobStatus.PENDING:
                job.status = JobStatus.RUNNING
                job.started_at = datetime.now(timezone.utc)
            db.session.commit()


def update_job_label(job_id: int, label: str) -> None:
    """Update job label in the database."""
    from models import Job, db

    app = _get_flask_app()
    with app.app_context():
        job = db.session.get(Job, job_id)
        if job:
            job.label = label
            db.session.commit()


def mark_job_completed(job_id: int, outputs: dict) -> None:
    """Mark a job as completed in the database."""
    from models import Artifact, ArtifactType, Job, JobStatus, Lecture, db

    app = _get_flask_app()
    with app.app_context():
        job = db.session.get(Job, job_id)
        if job:
            job.status = JobStatus.COMPLETED
            job.progress = 1.0
            job.completed_at = datetime.now(timezone.utc)

            # Create lecture and artifacts
            lecture = Lecture(
                user_id=job.user_id,
                job_id=job.id,
                title=job.label,
                source_id=outputs.get("source_id", "unknown"),
                date=datetime.now(timezone.utc),
            )
            db.session.add(lecture)
            db.session.flush()  # Get the lecture ID

            # Create artifacts
            artifact_mapping = {
                "pdf_path": ArtifactType.PDF_HANDOUT,
                "xlsx_path": ArtifactType.EXCEL_STUDY_TABLE,
                "vignette_path": ArtifactType.PDF_VIGNETTE,
            }

            for key, artifact_type in artifact_mapping.items():
                if key in outputs and outputs[key] and os.path.exists(outputs[key]):
                    file_path = outputs[key]
                    artifact = Artifact(
                        lecture_id=lecture.id,
                        artifact_type=artifact_type,
                        file_path=file_path,
                        file_name=os.path.basename(file_path),
                        file_size=os.path.getsize(file_path),
                    )
                    db.session.add(artifact)

            db.session.commit()


def mark_job_failed(job_id: int, error_message: str) -> None:
    """Mark a job as failed in the database."""
    from models import Job, JobStatus, db

    app = _get_flask_app()
    with app.app_context():
        job = db.session.get(Job, job_id)
        if job:
            job.status = JobStatus.FAILED
            job.error_message = error_message
            job.completed_at = datetime.now(timezone.utc)
            db.session.commit()


# Pipeline stage tasks


@celery_app.task(bind=True, base=PipelineTask, name="tasks.generate_context")
def generate_context_task(self, job_id: int, input_type: str, input_data: str) -> dict:
    """Generate processing context from input."""
    stage_name = "generate_context"

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Initializing")

    input_dict = json.loads(input_data)

    # Generate source_id based on input type
    if input_type == "panopto":
        source_id = input_dict.get("delivery_id", str(hash(input_data)))
    elif input_type == "url":
        source_id = str(hash(input_dict.get("url", "")))
    else:  # upload
        source_id = str(hash(input_dict.get("path", "")))

    ctx = TaskContext(
        source_id=source_id,
        input_type=input_type,
        input_data=input_dict,
        use_ai=True,
    )

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "Context created")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.download_video")
def download_video_task(self, data: dict) -> dict:
    """Download video if it doesn't exist."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])
    stage_name = "download_video"
    cache = CacheContext(ctx.source_id)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Checking video")

    # Check cache first
    cached = cache.get(stage_name)
    if cached and os.path.exists(cached.get("video_path", "")):
        ctx.video_path = cached["video_path"]
        self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
        update_job_progress(job_id, stage_name, 1.0, "Using cached video")
        return {"job_id": job_id, "context": ctx.to_dict()}

    video_path = ctx.get_video_path()

    if ctx.input_type == "upload":
        # For uploads, the file should already exist
        upload_path = ctx.input_data.get("path", "")
        if os.path.exists(upload_path):
            ctx.video_path = upload_path
            cache.set(stage_name, {"video_path": upload_path})
            update_job_progress(job_id, stage_name, 1.0, "Video ready")
            return {"job_id": job_id, "context": ctx.to_dict()}

    if ctx.input_type == "panopto":
        # Download from Panopto
        panopto_data = ctx.input_data
        _download_panopto_video(self, job_id, stage_name, panopto_data, video_path)
    else:
        # Download from URL
        video_url = ctx.input_data.get("url", "")
        if _is_m3u8_url(video_url):
            _download_m3u8_stream(self, job_id, stage_name, video_url, video_path)
        else:
            _download_regular_video(self, job_id, stage_name, video_url, video_path)

    ctx.video_path = video_path
    cache.set(stage_name, {"video_path": video_path})

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "Video downloaded")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.extract_captions")
def extract_captions_task(self, data: dict) -> dict:
    """Extract captions from the video."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])
    stage_name = "extract_captions"
    cache = CacheContext(ctx.source_id)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Extracting captions")

    # Check cache first
    cached = cache.get(stage_name)
    if cached is not None:
        ctx.captions = cached
        self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
        update_job_progress(job_id, stage_name, 1.0, "Using cached captions")
        return {"job_id": job_id, "context": ctx.to_dict()}

    update_job_progress(job_id, stage_name, 0.1, "Transcribing audio")

    # Run async transcription
    captions = run_async(generate_captions(ctx.get_video_path()))
    ctx.captions = [{"text": c.text, "timestamp": c.timestamp} for c in captions]

    cache.set(stage_name, ctx.captions)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "Captions extracted")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.match_frames")
def match_frames_task(self, data: dict) -> dict:
    """Match frames to captions based on structural similarity."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])
    stage_name = "match_frames"
    cache = CacheContext(ctx.source_id)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Matching frames")

    # Check cache first
    cached = cache.get(stage_name)
    if cached is not None:
        slides = cached
        if all(os.path.exists(s["image"]) for s in slides):
            ctx.slides = slides
            self.update_state(
                state="PROGRESS", meta={"stage": stage_name, "progress": 1.0}
            )
            update_job_progress(job_id, stage_name, 1.0, "Using cached slides")
            return {"job_id": job_id, "context": ctx.to_dict()}

    if not ctx.captions:
        ctx.slides = []
        return {"job_id": job_id, "context": ctx.to_dict()}

    captions = [Caption(**c) for c in ctx.captions]
    last_frame = None
    last_frame_gs = None
    cum_captions = []
    pairs = []

    stream = cv2.VideoCapture()
    stream.open(ctx.get_video_path())

    frame_path = os.path.join(FRAMES_DIR, ctx.source_id)
    os.makedirs(frame_path, exist_ok=True)

    for idx, cap in enumerate(captions):
        stream.set(cv2.CAP_PROP_POS_MSEC, cap.timestamp * 1_000 + 500)
        ret, frame = stream.read()
        if not ret:
            continue

        frame_gs = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        if last_frame is None:
            last_frame = frame
            last_frame_gs = frame_gs
            cum_captions.append(cap.text)
            continue

        similarity_result = ski.metrics.structural_similarity(
            last_frame_gs, frame_gs, full=False
        )
        score = (
            similarity_result
            if isinstance(similarity_result, (int, float))
            else similarity_result[0]
        )

        if score < 0.925 or (idx + 1) == len(captions):
            cap_full = " ".join(cum_captions)
            image_path = os.path.join(frame_path, f"{uuid4()}.png")
            cv2.imwrite(image_path, last_frame)

            pairs.append({"image": image_path, "caption": cap_full, "extra": None})
            last_frame = frame
            last_frame_gs = frame_gs
            cum_captions.clear()

            progress = (idx + 1) / len(captions)
            self.update_state(
                state="PROGRESS", meta={"stage": stage_name, "progress": progress}
            )
            update_job_progress(job_id, stage_name, progress * 0.9, "Matching slides")

        cum_captions.append(cap.text)

    stream.release()
    ctx.slides = pairs
    cache.set(stage_name, pairs)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "Frames matched")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.transform_slides_ai")
def transform_slides_ai_task(self, data: dict) -> dict:
    """Apply AI transformation to slides."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])
    stage_name = "transform_slides_with_ai"
    cache = CacheContext(ctx.source_id)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Transforming slides with AI")

    if not ctx.use_ai or not ctx.slides:
        return {"job_id": job_id, "context": ctx.to_dict()}

    # Check cache first
    cached = cache.get(stage_name)
    if cached is not None:
        slides = cached
        if all(os.path.exists(s["image"]) for s in slides):
            ctx.slides = slides
            self.update_state(
                state="PROGRESS", meta={"stage": stage_name, "progress": 1.0}
            )
            update_job_progress(job_id, stage_name, 1.0, "Using cached AI slides")
            return {"job_id": job_id, "context": ctx.to_dict()}

    output = []
    total = len(ctx.slides)

    for idx, slide in enumerate(ctx.slides):
        cleaned = run_async(clean_transcript(slide["caption"]))
        output.append(
            {
                "image": slide["image"],
                "caption": cleaned,
                "extra": slide.get("extra"),
            }
        )

        progress = (idx + 1) / total
        self.update_state(
            state="PROGRESS", meta={"stage": stage_name, "progress": progress}
        )
        update_job_progress(job_id, stage_name, progress * 0.9, "Cleaning transcript")

    ctx.slides = output
    cache.set(stage_name, output)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "Slides transformed")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.generate_output")
def generate_output_task(self, data: dict) -> dict:
    """Generate the PDF output."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])
    stage_name = "generate_output"
    cache = CacheContext(ctx.source_id)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Generating PDF")

    # Check cache first
    cached = cache.get(stage_name)
    if cached and os.path.exists(cached):
        ctx.outputs = ctx.outputs or {}
        ctx.outputs["pdf_path"] = cached
        self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
        update_job_progress(job_id, stage_name, 1.0, "Using cached PDF")
        return {"job_id": job_id, "context": ctx.to_dict()}

    # Convert slide dicts back to Slide namedtuples for template
    slides = [Slide(**s) for s in ctx.slides] if ctx.slides else []

    template_path = os.path.join(os.path.dirname(__file__), "templates")
    env = Environment(
        loader=FileSystemLoader(template_path), autoescape=select_autoescape()
    )
    template = env.get_template("template.html")
    html = template.render(pairs=slides)

    update_job_progress(job_id, stage_name, 0.3, "Generating title")
    title = run_async(generate_title(html))
    update_job_label(job_id, title)

    path = os.path.join(OUT_DIR, f"{title}.pdf")
    os.makedirs(OUT_DIR, exist_ok=True)

    update_job_progress(job_id, stage_name, 0.5, "Creating PDF")
    with open(path, "wb") as f:
        pisa_status = pisa.CreatePDF(html, dest=f)
        if hasattr(pisa_status, "err") and getattr(pisa_status, "err", None):
            raise ValueError("Error generating PDF")

    cache.set(stage_name, path)
    ctx.outputs = ctx.outputs or {}
    ctx.outputs["pdf_path"] = path
    ctx.outputs["source_id"] = ctx.source_id

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "PDF generated")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.compress_pdf")
def compress_pdf_task(self, data: dict) -> dict:
    """Compress the PDF using Ghostscript."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])
    stage_name = "compress_pdf"

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Compressing PDF")

    pdf_path = ctx.outputs.get("pdf_path") if ctx.outputs else None
    if not pdf_path or not os.path.exists(pdf_path):
        return {"job_id": job_id, "context": ctx.to_dict()}

    source_id = str(hash(pdf_path))
    cached = get_cached_result(source_id, stage_name)
    if cached and os.path.exists(cached):
        self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
        update_job_progress(job_id, stage_name, 1.0, "Using cached compressed PDF")
        return {"job_id": job_id, "context": ctx.to_dict()}

    with TemporaryDirectory() as temp_dir:
        output_path = os.path.join(temp_dir, f"compressed_{os.path.basename(pdf_path)}")

        gs_command = [
            "gs",
            "-sDEVICE=pdfwrite",
            "-dCompatibilityLevel=1.4",
            "-dPDFSETTINGS=/ebook",
            "-dNOPAUSE",
            "-dQUIET",
            "-dBATCH",
            f"-sOutputFile={output_path}",
            pdf_path,
        ]

        try:
            subprocess.run(gs_command, check=True)
            shutil.move(output_path, pdf_path)
        except (subprocess.CalledProcessError, FileNotFoundError) as e:
            print(f"Ghostscript compression failed: {e}")

    set_cached_result(source_id, stage_name, pdf_path)

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "PDF compressed")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.generate_spreadsheet")
def generate_spreadsheet_task(self, data: dict) -> dict:
    """Generate the Excel spreadsheet."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])
    stage_name = "generate_spreadsheet"

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Generating spreadsheet")

    if ctx.outputs is None:
        ctx.outputs = {}

    pdf_path = ctx.outputs.get("pdf_path")
    if not pdf_path or not os.path.exists(pdf_path):
        return {"job_id": job_id, "context": ctx.to_dict()}

    source_id = str(hash(pdf_path))
    cached = get_cached_result(source_id, stage_name)
    if cached:
        pdf_path, xlsx_path = cached
        if os.path.exists(xlsx_path):
            ctx.outputs["xlsx_path"] = xlsx_path
            self.update_state(
                state="PROGRESS", meta={"stage": stage_name, "progress": 1.0}
            )
            update_job_progress(job_id, stage_name, 1.0, "Using cached spreadsheet")
            return {"job_id": job_id, "context": ctx.to_dict()}

    update_job_progress(job_id, stage_name, 0.2, "Analyzing document")
    study_table = run_async(generate_spreadsheet_helper(pdf_path))

    if not study_table.rows:
        raise ValueError("No rows found in data")

    rows = [row.model_dump(by_alias=True) for row in study_table.rows]
    df = pd.DataFrame(rows)

    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    output_filename = os.path.join(OUT_DIR, f"{base_name}.xlsx")

    update_job_progress(job_id, stage_name, 0.6, "Writing Excel file")

    wb = Workbook()
    ws = wb.active
    assert ws is not None

    ws.title = "Study Table"

    # Write header row
    for col_num, column_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Write data rows
    for row_num, row_data in enumerate(rows, 2):
        for col_num, column_name in enumerate(df.columns, 1):
            cell_value = row_data.get(column_name, "")
            rich_text_value = parse_markdown_bold_to_rich_text(cell_value)
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = rich_text_value
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-adjust column widths
    for col_num in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    cell_length = min(len(str(cell.value)), 100)
                    max_length = max(max_length, cell_length)
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_filename)

    ctx.outputs["xlsx_path"] = output_filename
    set_cached_result(source_id, stage_name, (pdf_path, output_filename))

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "Spreadsheet generated")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.generate_vignette")
def generate_vignette_task(self, data: dict) -> dict:
    """Generate the vignette PDF."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])
    stage_name = "generate_vignette_pdf"

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 0})
    update_job_progress(job_id, stage_name, 0, "Generating vignette questions")

    if ctx.outputs is None:
        ctx.outputs = {}

    pdf_path = ctx.outputs.get("pdf_path")
    if not pdf_path or not os.path.exists(pdf_path):
        return {"job_id": job_id, "context": ctx.to_dict()}

    source_id = str(hash(pdf_path))
    cached = get_cached_result(source_id, stage_name)
    if cached:
        _, _, vignette_path = cached
        if os.path.exists(vignette_path):
            ctx.outputs["vignette_path"] = vignette_path
            self.update_state(
                state="PROGRESS", meta={"stage": stage_name, "progress": 1.0}
            )
            update_job_progress(job_id, stage_name, 1.0, "Using cached vignette")
            return {"job_id": job_id, "context": ctx.to_dict()}

    update_job_progress(job_id, stage_name, 0.2, "Generating questions")
    vignette_data = run_async(generate_vignette_questions(pdf_path))

    if not vignette_data.learning_objectives:
        raise ValueError("No learning objectives found")

    learning_objectives = [lo.model_dump() for lo in vignette_data.learning_objectives]

    template_path = os.path.join(os.path.dirname(__file__), "templates")
    env = Environment(
        loader=FileSystemLoader(template_path), autoescape=select_autoescape()
    )
    template = env.get_template("vignette.html")
    html = template.render(learning_objectives=learning_objectives)

    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    vignette_pdf_path = os.path.join(OUT_DIR, f"{base_name} - Vignette Questions.pdf")

    update_job_progress(job_id, stage_name, 0.7, "Creating vignette PDF")

    with open(vignette_pdf_path, "wb") as f:
        pisa_status = pisa.CreatePDF(html, dest=f)
        if hasattr(pisa_status, "err") and getattr(pisa_status, "err", None):
            raise ValueError("Error generating vignette PDF")

    ctx.outputs["vignette_path"] = vignette_pdf_path
    xlsx_path = ctx.outputs.get("xlsx_path", "")
    set_cached_result(source_id, stage_name, (pdf_path, xlsx_path, vignette_pdf_path))

    self.update_state(state="PROGRESS", meta={"stage": stage_name, "progress": 1.0})
    update_job_progress(job_id, stage_name, 1.0, "Vignette generated")

    return {"job_id": job_id, "context": ctx.to_dict()}


@celery_app.task(bind=True, base=PipelineTask, name="tasks.finalize_job")
def finalize_job_task(self, data: dict) -> dict:
    """Finalize the job and create database records."""
    job_id = data["job_id"]
    ctx = TaskContext.from_dict(data["context"])

    outputs = ctx.outputs or {}
    outputs["source_id"] = ctx.source_id

    mark_job_completed(job_id, outputs)

    return {"job_id": job_id, "status": "completed", "outputs": outputs}


# Helper functions for video download


def _is_m3u8_url(url: str) -> bool:
    """Check if a URL points to an M3U8 file."""
    return url.endswith(".m3u8") or "m3u8" in url


def _download_regular_video(
    task, job_id: int, stage_name: str, video_url: str, video_path: str
) -> None:
    """Download a regular video file."""

    def report_progress(count, block_size, total_size):
        if total_size > 0:
            progress = count * block_size / total_size
            task.update_state(
                state="PROGRESS", meta={"stage": stage_name, "progress": progress}
            )
            update_job_progress(job_id, stage_name, progress * 0.9, "Downloading")

    opener = urllib.request.build_opener()
    opener.addheaders = [("Range", "bytes=0-")]
    urllib.request.install_opener(opener)

    urllib.request.urlretrieve(video_url, video_path, reporthook=report_progress)


def _download_m3u8_stream(
    task, job_id: int, stage_name: str, video_url: str, video_path: str
) -> None:
    """Download and combine M3U8 stream segments."""
    update_job_progress(job_id, stage_name, 0.1, "Parsing playlist")

    playlist = m3u8.load(video_url)

    if playlist.is_variant:
        if playlist.playlists:
            best_playlist = min(
                playlist.playlists,
                key=lambda p: p.stream_info.bandwidth if p.stream_info.bandwidth else 0,
            )
            stream_url = urljoin(video_url, best_playlist.uri)
            playlist = m3u8.load(stream_url)
        else:
            raise ValueError("No streams found in variant playlist")

    segments = playlist.segments
    total_segments = len(segments)

    if total_segments == 0:
        raise ValueError("No segments found in playlist")

    with tempfile.TemporaryDirectory() as temp_dir:
        segment_files = []

        for i, segment in enumerate(segments):
            segment_url = urljoin(playlist.base_uri or video_url, segment.uri)
            segment_path = os.path.join(temp_dir, f"segment_{i:04d}.ts")

            for attempt in range(3):
                try:
                    urllib.request.urlretrieve(segment_url, segment_path)
                    if os.path.getsize(segment_path) > 0:
                        break
                except Exception as e:
                    if attempt == 2:
                        raise ValueError(f"Failed to download segment {i}: {e}")

            segment_files.append(segment_path)
            progress = (i + 1) / total_segments * 0.8
            update_job_progress(job_id, stage_name, progress, "Downloading segments")

        update_job_progress(job_id, stage_name, 0.85, "Combining segments")

        concat_file = os.path.join(temp_dir, "segments.txt")
        with open(concat_file, "w") as f:
            for segment_file in segment_files:
                f.write(f"file '{segment_file}'\n")

        result = subprocess.run(
            [
                "ffmpeg",
                "-f",
                "concat",
                "-safe",
                "0",
                "-i",
                concat_file,
                "-c",
                "copy",
                "-y",
                video_path,
            ],
            capture_output=True,
            text=True,
        )

        if result.returncode != 0:
            with open(video_path, "wb") as outfile:
                for segment_file in segment_files:
                    with open(segment_file, "rb") as infile:
                        outfile.write(infile.read())


def _download_panopto_video(
    task, job_id: int, stage_name: str, panopto_data: dict, video_path: str
) -> None:
    """Download video from Panopto."""
    base = panopto_data["base"]
    cookie = panopto_data["cookie"]
    delivery_id = panopto_data["delivery_id"]

    update_job_progress(job_id, stage_name, 0.1, "Getting Panopto info")

    delivery_info = fetch(
        base,
        cookie,
        "Panopto/Pages/Viewer/DeliveryInfo.aspx",
        {
            "deliveryId": delivery_id,
            "responseType": "json",
            "getCaptions": "false",
            "language": "0",
        },
    )

    vidurl = delivery_info["Delivery"]["PodcastStreams"][0]["StreamUrl"]

    if _is_m3u8_url(vidurl):
        _download_m3u8_stream(task, job_id, stage_name, vidurl, video_path)
    else:
        _download_regular_video(task, job_id, stage_name, vidurl, video_path)


def create_pipeline_chain(job_id: int, input_type: str, input_data: str):
    """Create a Celery chain for the full pipeline."""
    return chain(
        generate_context_task.s(job_id, input_type, input_data),
        download_video_task.s(),
        extract_captions_task.s(),
        match_frames_task.s(),
        transform_slides_ai_task.s(),
        generate_output_task.s(),
        compress_pdf_task.s(),
        generate_spreadsheet_task.s(),
        generate_vignette_task.s(),
        finalize_job_task.s(),
    )


def start_pipeline(job_id: int, input_type: str, input_data: str) -> str:
    """Start the pipeline and return the Celery task ID."""
    pipeline = create_pipeline_chain(job_id, input_type, input_data)
    result = pipeline.apply_async()
    return result.id
    return result.id
