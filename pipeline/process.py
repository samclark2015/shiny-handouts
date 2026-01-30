import asyncio
import os
import re
import shutil
import subprocess
import tempfile
import urllib.request
from collections import namedtuple
from dataclasses import dataclass
from tempfile import TemporaryDirectory
from typing import Callable, cast
from urllib.parse import urljoin
from uuid import uuid4

import cv2
import m3u8
import pandas as pd
import skimage as ski
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

from pipeline.ai import (
    clean_transcript,
    generate_captions,
    generate_spreadsheet_helper,
    generate_title,
    generate_vignette_questions,
)
from pipeline.cache import get_cached_result, set_cached_result
from pipeline.helpers import Caption, Slide, fetch

from .pipeline import Pipeline, PipelineFailure, Progress

in_dir = os.path.join("data", "input")
out_dir = os.path.join("data", "output")


def parse_markdown_bold_to_rich_text(text: str) -> CellRichText | str:
    """
    Parse Markdown bold syntax (**text**) and convert to Excel rich text.
    Returns CellRichText if bold markers are found, otherwise returns the original string.
    """
    if not text or not isinstance(text, str):
        return text or ""

    # Pattern to match **bold** text
    pattern = r"\*\*(.+?)\*\*"

    # Check if there are any bold markers
    if not re.search(pattern, text):
        return text

    # Split text into parts (bold and non-bold)
    parts = []
    last_end = 0

    for match in re.finditer(pattern, text):
        # Add non-bold text before this match
        if match.start() > last_end:
            non_bold_text = text[last_end : match.start()]
            if non_bold_text:
                parts.append(non_bold_text)

        # Add bold text
        bold_text = match.group(1)
        if bold_text:
            bold_font = InlineFont(b=True)
            parts.append(TextBlock(bold_font, bold_text))

        last_end = match.end()

    # Add any remaining non-bold text after the last match
    if last_end < len(text):
        remaining_text = text[last_end:]
        if remaining_text:
            parts.append(remaining_text)

    # Return CellRichText if we have parts, otherwise original text
    if parts:
        return CellRichText(parts)
    return text


PanoptoInput = namedtuple("PanoptoInput", ["base", "cookie", "delivery_id"])

ProcessingInput = str | PanoptoInput

ENABLE_AI = True  # Set to False to disable AI processing


@dataclass
class ProcessingContext:
    """Context object that flows through the pipeline stages."""

    pipeline: Pipeline
    source_id: str
    source: str | PanoptoInput
    use_ai: bool
    captions: list[Caption] | None = None
    slides: list[Slide] | None = None

    @property
    def video_path(self) -> str:
        if isinstance(self.source, str) and os.path.exists(self.source):
            return self.source
        return os.path.join(in_dir, f"video_{self.source_id}.mp4")

    def get_cached(self, stage_name: str):
        """Get cached result for a stage."""
        return get_cached_result(self.source_id, stage_name)

    def set_cached(self, stage_name: str, result):
        """Cache result for a stage."""
        set_cached_result(self.source_id, stage_name, result)


# Utility functions for video download
def _is_m3u8_url(url: str) -> bool:
    """Check if a URL points to an M3U8 file."""
    return url.endswith(".m3u8") or "m3u8" in url


def _download_m3u8_stream(ctx: ProcessingContext, video_url: str) -> None:
    """Download and combine M3U8 stream segments into a single video file."""
    ctx.pipeline.report_progress("Parsing playlist")

    try:
        # Parse the m3u8 playlist
        playlist = m3u8.load(video_url)

        if playlist.is_variant:
            # Select the highest quality stream or first available
            if playlist.playlists:
                # Sort by bandwidth (highest first) and select the best quality
                best_playlist = min(
                    playlist.playlists,
                    key=lambda p: p.stream_info.bandwidth
                    if p.stream_info.bandwidth
                    else 0,
                )
                stream_url = urljoin(video_url, best_playlist.uri)
                playlist = m3u8.load(stream_url)
            else:
                raise PipelineFailure("No streams found in variant playlist")

        # Download and combine segments
        segments = playlist.segments
        total_segments = len(segments)

        if total_segments == 0:
            raise PipelineFailure("No segments found in playlist")

        # Create a temporary directory for segments
        with tempfile.TemporaryDirectory() as temp_dir:
            segment_files = []

            # Download each segment with retry logic
            for i, segment in enumerate(segments):
                segment_url = urljoin(playlist.base_uri or video_url, segment.uri)
                segment_path = os.path.join(temp_dir, f"segment_{i:04d}.ts")

                # Retry download up to 3 times
                for attempt in range(3):
                    try:
                        urllib.request.urlretrieve(segment_url, segment_path)
                        # Verify the segment was downloaded completely
                        if os.path.getsize(segment_path) > 0:
                            break
                    except Exception as e:
                        if attempt == 2:  # Last attempt
                            raise ValueError(
                                f"Failed to download segment {i} after 3 attempts: {e}"
                            )
                        continue

                segment_files.append(segment_path)

                # Update progress
                ctx.pipeline.report_progress(
                    "Downloading video segments", (i + 1) / total_segments
                )

            # Use ffmpeg to properly concatenate segments instead of binary concatenation
            ctx.pipeline.report_progress("Combining segments")

            # Create a file list for ffmpeg
            concat_file = os.path.join(temp_dir, "segments.txt")
            with open(concat_file, "w") as f:
                for segment_file in segment_files:
                    f.write(f"file '{segment_file}'\n")

            # Use ffmpeg to concatenate properly
            result = subprocess.run(
                [
                    "ffmpeg",
                    "-f",
                    "concat",
                    "-safe",
                    "0",
                    "-i",
                    concat_file,
                    "-c",
                    "copy",
                    "-y",
                    ctx.video_path,
                ],
                capture_output=True,
                text=True,
            )

            if result.returncode != 0:
                # Fallback to binary concatenation if ffmpeg fails
                with open(ctx.video_path, "wb") as outfile:
                    for segment_file in segment_files:
                        with open(segment_file, "rb") as infile:
                            outfile.write(infile.read())

    except urllib.request.HTTPError as e:
        raise PipelineFailure(
            f"HTTP Error while downloading m3u8 stream: {str(e)} @ {e.url}"
        )
    except Exception as e:
        raise PipelineFailure(f"Failed to download m3u8 stream: {str(e)}")


def _download_regular_video(
    ctx: ProcessingContext, video_url: str, use_range_header: bool = True
) -> ProcessingContext:
    """Download a regular video file."""
    if use_range_header:
        opener = urllib.request.build_opener()
        opener.addheaders = [("Range", "bytes=0-")]
        urllib.request.install_opener(opener)

    urllib.request.urlretrieve(
        video_url,
        ctx.video_path,
        reporthook=lambda count, bs, ts: ctx.pipeline.report_progress(
            "Downloading", count * bs / ts
        ),
    )
    return ctx


# Panopto-specific functions
def _get_delivery_info(
    base: str, cookie: str, delivery_id: str, captions: bool = False
):
    """Get delivery info from Panopto."""
    url = "Panopto/Pages/Viewer/DeliveryInfo.aspx"
    data = fetch(
        base,
        cookie,
        url,
        {
            "deliveryId": delivery_id,
            "responseType": "json",
            "getCaptions": "true" if captions else "false",
            "language": "0",
        },
    )
    return data


def _download_panopto_video(ctx: ProcessingContext) -> ProcessingContext:
    """Download video from Panopto if it doesn't exist."""
    # Get the video URL from Panopto
    panopto = cast(PanoptoInput, ctx.source)
    delivery_info = _get_delivery_info(
        panopto.base, panopto.cookie, panopto.delivery_id
    )
    vidurl = delivery_info["Delivery"]["PodcastStreams"][0]["StreamUrl"]

    # Check if the URL points to an m3u8 file
    if _is_m3u8_url(vidurl):
        # Use m3u8 library to parse and download stream
        _download_m3u8_stream(ctx, vidurl)
    else:
        # Regular video file download
        _download_regular_video(ctx, vidurl, False)

    print("Downloaded video to", ctx.video_path)
    return ctx


# Pipeline stage functions


def generate_context(pipeline: Pipeline, input: ProcessingInput) -> ProcessingContext:
    """Generate processing context from input."""
    source_id = (
        input.delivery_id if isinstance(input, PanoptoInput) else str(hash(input))
    )
    return ProcessingContext(
        pipeline,
        source_id,
        input,
        ENABLE_AI,
    )


def download_video(pipeline: Pipeline, ctx: ProcessingContext) -> ProcessingContext:
    """Download video if it doesn't exist."""
    stage_name = "download_video"

    # Check cache first
    cached = ctx.get_cached(stage_name)
    if cached is not None and os.path.exists(cached.get("video_path", "")):
        pipeline.report_progress("Using cached video", 1.0)
        return ctx

    if (
        isinstance(ctx.source, str)
        and os.path.exists(ctx.source)
        and ctx.source == ctx.video_path
    ):
        ctx.set_cached(stage_name, {"video_path": ctx.video_path})
        return ctx

    if not os.path.exists(ctx.video_path):
        # Check if the URL points to an m3u8 file
        if isinstance(ctx.source, PanoptoInput):
            result = _download_panopto_video(ctx)
        else:
            result = _download_regular_video(ctx, ctx.source)
        ctx.set_cached(stage_name, {"video_path": ctx.video_path})
        return result

    ctx.set_cached(stage_name, {"video_path": ctx.video_path})
    return ctx


async def extract_captions(
    pipeline: Pipeline, ctx: ProcessingContext
) -> ProcessingContext:
    """Extract captions from the video."""
    stage_name = "extract_captions"

    # Check cache first
    cached = ctx.get_cached(stage_name)
    if cached is not None:
        pipeline.report_progress("Using cached captions", 1.0)
        ctx.captions = [Caption(**c) for c in cached]
        return ctx

    ctx.pipeline.report_progress("Transcribing")
    ctx.captions = await generate_captions(ctx.video_path)

    # Cache the captions (convert namedtuples to dicts for serialization)
    ctx.set_cached(stage_name, [c._asdict() for c in ctx.captions])
    return ctx


def match_frames(pipeline: Pipeline, ctx: ProcessingContext) -> ProcessingContext:
    """Match frames to captions based on structural similarity."""
    stage_name = "match_frames"

    # Check cache first
    cached = ctx.get_cached(stage_name)
    if cached is not None:
        # Verify cached frame images still exist
        slides = [Slide(**s) for s in cached]
        if all(os.path.exists(s.image) for s in slides):
            pipeline.report_progress("Using cached slides", 1.0)
            ctx.slides = slides
            return ctx

    if not ctx.captions:
        ctx.slides = []
        return ctx

    last_frame = None
    last_frame_gs = None
    cum_captions = []

    pairs: list[Slide] = []

    stream = cv2.VideoCapture()
    stream.open(ctx.video_path)

    frame_path = os.path.join("data", "frames", ctx.source_id)
    os.makedirs(frame_path, exist_ok=True)

    for idx, cap in enumerate(ctx.captions):
        stream.set(cv2.CAP_PROP_POS_MSEC, cap.timestamp * 1_000 + 500)
        ret, frame = stream.read()
        if not ret:
            raise ValueError("Could not read frame")

        frame_gs = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        if last_frame is None:
            last_frame = frame
            last_frame_gs = frame_gs
            cum_captions.append(cap.text)
            continue

        similarity_result = ski.metrics.structural_similarity(
            last_frame_gs, frame_gs, full=False
        )
        # Handle both single score and tuple return types
        score = (
            similarity_result
            if isinstance(similarity_result, (int, float))
            else similarity_result[0]
        )

        if score < 0.925 or (idx + 1) == len(ctx.captions):
            cap_full = " ".join(cum_captions)
            image_path = os.path.join(frame_path, f"{uuid4()}.png")
            cv2.imwrite(image_path, last_frame)

            pairs.append(Slide(image_path, cap_full, None))
            last_frame = frame
            last_frame_gs = frame_gs
            cum_captions.clear()
            ctx.pipeline.report_progress(
                "Matching Slides", (idx + 1) / len(ctx.captions)
            )
        cum_captions.append(cap.text)

    stream.release()
    ctx.slides = pairs

    # Cache the slides (convert namedtuples to dicts for serialization)
    ctx.set_cached(stage_name, [s._asdict() for s in pairs])
    return ctx


async def transform_slides_with_ai(
    pipeline: Pipeline, ctx: ProcessingContext
) -> ProcessingContext:
    """Apply AI transformation to slides if enabled."""
    stage_name = "transform_slides_with_ai"

    if not ctx.use_ai or not ctx.slides:
        return ctx

    # Check cache first
    cached = ctx.get_cached(stage_name)
    if cached is not None:
        # Verify cached frame images still exist
        slides = [Slide(**s) for s in cached]
        if all(os.path.exists(s.image) for s in slides):
            pipeline.report_progress("Using cached AI-transformed slides", 1.0)
            ctx.slides = slides
            return ctx

    async def transform_slide(slide: Slide) -> Slide:
        """Transform a single slide using AI."""
        cleaned = await clean_transcript(slide.caption)
        keypoints = None
        # cleaned, keypoints = await asyncio.gather(clean_transcript(slide.caption), gen_keypoints(slide.caption, slide.image))
        return Slide(slide.image, cleaned, keypoints)

    output: list[Slide] = []
    for idx, slide in enumerate(ctx.slides):
        ctx.pipeline.report_progress(
            "Cleaning Transcript with AI", (idx + 1) / len(ctx.slides)
        )
        output.append(await transform_slide(slide))
    ctx.slides = output

    # Cache the transformed slides
    ctx.set_cached(stage_name, [s._asdict() for s in output])
    return ctx


def generate_pdf_output(ctx: ProcessingContext, html: str, path: str) -> None:
    """Generate PDF from HTML."""
    with open(path, "wb") as f:
        pisa_status = pisa.CreatePDF(html, dest=f)
        # Check for errors if available
        if hasattr(pisa_status, "err") and getattr(pisa_status, "err", None):
            raise ValueError("Error generating PDF")


async def generate_output(pipeline: Pipeline, ctx: ProcessingContext) -> str:
    """Generate the final PDF output."""
    stage_name = "generate_output"

    # Check cache first
    cached = ctx.get_cached(stage_name)
    if cached is not None and os.path.exists(cached):
        pipeline.report_progress("Using cached PDF", 1.0)
        return cached

    pipeline.report_progress("Generating PDF", 0)
    template_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "templates"
    )
    env = Environment(
        loader=FileSystemLoader(template_path), autoescape=select_autoescape()
    )
    template = env.get_template("template.html")

    html = template.render(pairs=ctx.slides)

    title = await generate_title(html)
    path = os.path.join(out_dir, f"{title}.pdf")
    os.makedirs(out_dir, exist_ok=True)

    # Run PDF generation in executor since it's CPU-bound
    await asyncio.get_event_loop().run_in_executor(
        None, generate_pdf_output, ctx, html, path
    )
    pipeline.report_progress("Generating PDF", 1.0)

    # Cache the output path
    ctx.set_cached(stage_name, path)
    return path


def compress_pdf(pipeline: Pipeline, input_path: str, quality="ebook") -> str:
    """
    Compresses a PDF file using Ghostscript.

    Args:
        input_path (str): Path to the input PDF file.
        output_path (str): Path where the compressed PDF will be saved.
        quality (str): Quality setting for compression ('screen', 'ebook',
                       'printer', 'prepress'). 'ebook' is a good balance.
    """
    stage_name = "compress_pdf"
    source_id = str(hash(input_path))

    # Check cache first - if the file exists and matches cached path, skip compression
    cached = get_cached_result(source_id, stage_name)
    if cached is not None and os.path.exists(cached):
        pipeline.report_progress("Using cached compressed PDF", 1.0)
        return cached

    with TemporaryDirectory() as temp_dir:
        output_path = os.path.join(
            temp_dir, f"compressed_{os.path.basename(input_path)}"
        )

        gs_command = [
            "gs",
            "-sDEVICE=pdfwrite",
            "-dCompatibilityLevel=1.4",
            f"-dPDFSETTINGS=/{quality}",
            "-dNOPAUSE",
            "-dQUIET",
            "-dBATCH",
            f"-sOutputFile={output_path}",
            input_path,
        ]

        pipeline.report_progress("Compressing PDF", 0)

        try:
            subprocess.run(gs_command, check=True)
        except subprocess.CalledProcessError as e:
            print(f"Error during Ghostscript execution: {e}")
            return input_path
        except FileNotFoundError:
            print("Ghostscript not found. Ensure it is installed and in your PATH.")
            return input_path

        shutil.move(output_path, input_path)
        pipeline.report_progress("Compressing PDF", 1.0)

    # Cache the result
    set_cached_result(source_id, stage_name, input_path)
    return input_path


async def generate_spreadsheet(pipeline: Pipeline, filename: str) -> tuple[str, str]:
    stage_name = "generate_spreadsheet"
    source_id = str(hash(filename))

    # Check cache first
    cached = get_cached_result(source_id, stage_name)
    if cached is not None:
        pdf_path, xlsx_path = cached
        if os.path.exists(pdf_path) and os.path.exists(xlsx_path):
            pipeline.report_progress("Using cached Excel sheet", 1.0)
            return cached

    pipeline.report_progress("Generating Excel Sheet", 0)

    study_table = await generate_spreadsheet_helper(filename)

    # Extract rows from the Pydantic model
    if not study_table.rows:
        raise ValueError("No rows found in data")

    # Convert Pydantic models to dicts for DataFrame/Excel processing
    rows = [row.model_dump(by_alias=True) for row in study_table.rows]

    # Convert to DataFrame
    df = pd.DataFrame(rows)

    # Generate output filename
    base_name = os.path.splitext(os.path.basename(filename))[0]
    output_filename = os.path.join(out_dir, f"{base_name}.xlsx")
    os.makedirs(out_dir, exist_ok=True)

    # Write to Excel with rich text support for Markdown bold
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Worksheet not found in the workbook"
    ws.title = "Study Table"

    # Write header row
    for col_num, column_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Write data rows with Markdown bold parsing
    for row_num, row_data in enumerate(rows, 2):  # Start from row 2 (after header)
        for col_num, column_name in enumerate(df.columns, 1):
            cell_value = row_data.get(column_name, "")
            rich_text_value = parse_markdown_bold_to_rich_text(cell_value)
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = rich_text_value
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-adjust column widths for study table
    for col_num in range(1, len(df.columns) + 1):
        column_letter = get_column_letter(col_num)

        # Calculate max length in column
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    # Limit to reasonable width
                    cell_length = min(len(str(cell.value)), 100)
                    max_length = max(max_length, cell_length)
            except Exception:
                pass

        # Set column width (add a bit of padding)
        adjusted_width = min(max_length + 2, 80)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the formatted workbook
    wb.save(output_filename)
    pipeline.report_progress("Generating Excel Sheet", 1.0)

    # Cache the result
    result = (filename, output_filename)
    set_cached_result(source_id, stage_name, result)
    return result


async def generate_vignette_pdf(
    pipeline: Pipeline, inputs: tuple[str, str]
) -> tuple[str, str, str]:
    """Generate a PDF with vignette questions for each learning objective."""
    pdf_filename, xlsx_filename = inputs
    stage_name = "generate_vignette_pdf"
    source_id = str(hash(pdf_filename))

    # Check cache first
    cached = get_cached_result(source_id, stage_name)
    if cached is not None:
        pdf_path, xlsx_path, vignette_path = cached
        if all(os.path.exists(p) for p in [pdf_path, xlsx_path, vignette_path]):
            pipeline.report_progress("Using cached vignette PDF", 1.0)
            return cached

    pipeline.report_progress("Generating Vignette Questions", 0)

    # Generate vignette questions from the PDF
    vignette_data = await generate_vignette_questions(pdf_filename)

    pipeline.report_progress("Generating Vignette Questions", 0.5)

    # Extract learning objectives from the Pydantic model
    if not vignette_data.learning_objectives:
        raise PipelineFailure("No learning objectives found in the lecture")

    # Convert to dicts for template rendering
    learning_objectives = [lo.model_dump() for lo in vignette_data.learning_objectives]

    # Render the HTML template
    template_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "templates"
    )
    env = Environment(
        loader=FileSystemLoader(template_path), autoescape=select_autoescape()
    )
    template = env.get_template("vignette.html")

    html = template.render(learning_objectives=learning_objectives)

    # Generate output filename
    base_name = os.path.splitext(os.path.basename(pdf_filename))[0]
    vignette_pdf_path = os.path.join(out_dir, f"{base_name} - Vignette Questions.pdf")
    os.makedirs(out_dir, exist_ok=True)

    pipeline.report_progress("Generating Vignette PDF", 0.7)

    # Generate the PDF
    with open(vignette_pdf_path, "wb") as f:
        pisa_status = pisa.CreatePDF(html, dest=f)
        if hasattr(pisa_status, "err") and getattr(pisa_status, "err", None):
            raise PipelineFailure("Error generating vignette PDF")

    pipeline.report_progress("Generating Vignette PDF", 1.0)

    # Cache the result
    result = (pdf_filename, xlsx_filename, vignette_pdf_path)
    set_cached_result(source_id, stage_name, result)
    return result


def create_pipeline(
    callback: Callable[[Pipeline, Progress], None],
) -> Pipeline[ProcessingInput, tuple[str, str, str]]:
    pipeline = (
        Pipeline[ProcessingInput](callback)
        .add_stage(generate_context)
        .add_stage(download_video)
        .add_stage(extract_captions)
        .add_stage(match_frames)
        .add_stage(transform_slides_with_ai)
        .add_stage(generate_output)
        .add_stage(compress_pdf)
        .add_stage(generate_spreadsheet)
        .add_stage(generate_vignette_pdf)
    )
    return pipeline
