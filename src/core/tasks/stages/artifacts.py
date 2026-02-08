"""
Artifact generation stage tasks (spreadsheet, vignette, mindmap).
"""

import asyncio
import logging
import os
from contextlib import AsyncExitStack
from tempfile import NamedTemporaryFile, TemporaryDirectory

import pandas as pd
from django.conf import settings
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

from core.storage import is_s3_enabled, temp_download, upload_job_file
from core.tasks.config import broker
from core.tasks.context import TaskContext
from core.tasks.db import create_artifact
from core.tasks.progress import update_job_progress
from pipeline.ai import generate_mindmap, generate_spreadsheet_helper, generate_vignette_questions
from pipeline.helpers import parse_markdown_bold_to_rich_text


async def _prepare_images_for_llm(
    slides: list[dict] | None,
    temp_dir: str,
) -> tuple[list[dict], dict[str, str]]:
    """Download slide images and prepare them for LLM.

    Args:
        slides: List of slide dicts from match_frames_task with 'image', 'caption', 'extra' keys
        temp_dir: Temporary directory for downloaded images

    Returns:
        Tuple of (images_for_llm, image_id_to_local_path mapping)
    """
    if not slides:
        return [], {}

    images_for_llm = []
    id_to_path: dict[str, str] = {}

    for idx, slide_data in enumerate(slides):
        img_id = f"img_{idx:03d}"
        img_path = slide_data.get("image", "")

        if not img_path:
            continue

        try:
            # Download from S3 if needed
            if is_s3_enabled():
                local_path = os.path.join(temp_dir, f"{img_id}.jpg")
                from core.storage import S3Storage, get_storage_config

                config = get_storage_config()
                storage = S3Storage(config)
                await storage.download_file(img_path, local_path)
            else:
                local_path = img_path

            # Create image data for LLM with local path and slide metadata
            img_for_llm = {
                "path": local_path,
                "caption": slide_data.get("caption", ""),
                "extra": slide_data.get("extra"),
            }
            images_for_llm.append(img_for_llm)
            id_to_path[img_id] = local_path

        except Exception as e:
            logging.warning(f"Failed to prepare image {img_path}: {e}")
            continue

    return images_for_llm, id_to_path


def _insert_images_into_excel(
    ws,
    study_table_rows: list[dict],
    image_id_to_path: dict[str, str],
    first_image_column: int,
    start_row: int = 2,
    max_image_height: int = 100,
):
    """Insert images into Excel worksheet based on LLM selections.

    Each image gets its own column (Image 1, Image 2, etc.)

    Args:
        ws: openpyxl worksheet
        study_table_rows: Rows from the study table with image_ids
        image_id_to_path: Mapping from image IDs to local file paths
        first_image_column: Column number for the first image column
        start_row: First data row (after headers)
        max_image_height: Maximum height for inserted images in pixels
    """
    from PIL import Image as PILImage

    for row_idx, row_data in enumerate(study_table_rows):
        excel_row = start_row + row_idx
        image_ids = row_data.get("image_ids", [])

        if not image_ids:
            continue

        row_height_points = 15  # Default row height
        images_inserted = 0

        for img_idx, img_id in enumerate(image_ids):
            img_path = image_id_to_path.get(img_id)
            if not img_path or not os.path.exists(img_path):
                continue

            try:
                # Get image dimensions
                with PILImage.open(img_path) as pil_img:
                    orig_width, orig_height = pil_img.size

                # Calculate scaled dimensions (maintain aspect ratio)
                if orig_height > max_image_height:
                    scale = max_image_height / orig_height
                    new_width = int(orig_width * scale)
                    new_height = max_image_height
                else:
                    new_width = orig_width
                    new_height = orig_height

                # Create openpyxl image
                xl_img = XLImage(img_path)
                xl_img.width = new_width
                xl_img.height = new_height

                # Position the image in its respective column
                image_column = first_image_column + img_idx
                col_letter = get_column_letter(image_column)
                anchor = f"{col_letter}{excel_row}"
                xl_img.anchor = anchor

                # Add to worksheet
                ws.add_image(xl_img)
                images_inserted += 1

                # Track max height needed for this row
                # Convert pixels to points (1 point = 1.333 pixels approx)
                height_points = new_height * 0.75
                row_height_points = max(row_height_points, height_points)

            except Exception as e:
                logging.warning(f"Failed to insert image {img_id}: {e}")
                continue

        # Set row height to accommodate images
        if images_inserted > 0:
            ws.row_dimensions[excel_row].height = row_height_points + 10


@broker.task
async def generate_spreadsheet_artifact_task(data: dict) -> dict:
    """Generate Excel spreadsheet artifact as a distributed task.

    Returns dict with xlsx_path if successful, empty dict if skipped/failed.
    """
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    user_id = ctx.user_id

    if not ctx.enable_excel:
        return {}

    if ctx.outputs is None:
        return {}

    pdf_path = ctx.outputs.get("pdf_path")
    if not pdf_path:
        return {}

    try:
        # Create temp directory for image processing
        with TemporaryDirectory() as temp_dir:
            # Prepare slide images for LLM (from match_frames_task)
            images_for_llm, image_id_to_path = await _prepare_images_for_llm(ctx.slides, temp_dir)

            # Download PDF if on S3 for AI processing
            async with temp_download(pdf_path) as local_pdf:
                study_table = await generate_spreadsheet_helper(
                    local_pdf,
                    custom_prompt=ctx.spreadsheet_prompt,
                    custom_columns=ctx.spreadsheet_columns,
                    extracted_images=images_for_llm if images_for_llm else None,
                    user_id=user_id,
                    job_id=job_id,
                )

            if not study_table.rows:
                return {}

            # Determine maximum number of images in any row
            max_images = 0
            if images_for_llm:
                for row in study_table.rows:
                    image_ids = row.get("image_ids", [])
                    if isinstance(image_ids, list):
                        max_images = max(max_images, len(image_ids))

            # Build column list (text columns only, images added separately)
            text_columns = [col for col in study_table.rows[0].keys() if col != "image_ids"]

            # Create image column headers
            image_column_names = [f"Image {i + 1}" for i in range(max_images)]
            display_columns = text_columns + image_column_names

            df = pd.DataFrame(
                [{k: v for k, v in row.items() if k != "image_ids"} for row in study_table.rows]
            )

            # Get base name from the PDF filename
            pdf_filename = ctx.outputs.get("pdf_filename", os.path.basename(pdf_path))
            base_name = os.path.splitext(pdf_filename)[0]
            output_filename = f"{base_name}.xlsx"

            # Style constants
            HEADER_BG_COLOR = "D3D3D3"
            CELL_BG_COLOR = "ADD8E6"
            SECTION_HEADER_BG_COLOR = "6CB4E8"
            BORDER_COLOR = "000000"

            thin_border = Border(
                left=Side(style="thin", color=BORDER_COLOR),
                right=Side(style="thin", color=BORDER_COLOR),
                top=Side(style="thin", color=BORDER_COLOR),
                bottom=Side(style="thin", color=BORDER_COLOR),
            )

            wb = Workbook()
            ws = wb.active
            assert ws is not None

            ws.title = "Study Table"

            # Write header row
            for col_num, column_name in enumerate(display_columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column_name)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid"
                )
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

            # Track first image column for later insertion
            first_image_column = len(text_columns) + 1 if max_images > 0 else None

            # Write data rows
            for row_num, row_data in enumerate(study_table.rows, 2):
                first_cell_value = row_data.get(text_columns[0], "") if text_columns else ""
                is_section_header = all(
                    row_data.get(c, "") == "" or row_data.get(c, "") == first_cell_value
                    for c in text_columns
                )

                for col_num, column_name in enumerate(display_columns, 1):
                    # Check if this is an image column
                    if column_name.startswith("Image "):
                        # Leave empty - images will be inserted later
                        cell = ws.cell(row=row_num, column=col_num, value="")
                        cell.fill = PatternFill(
                            start_color=CELL_BG_COLOR, end_color=CELL_BG_COLOR, fill_type="solid"
                        )
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        cell.border = thin_border
                        continue

                    cell_value = row_data.get(column_name, "")

                    if is_section_header and col_num == 1:
                        cell = ws.cell(row=row_num, column=col_num, value=first_cell_value)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color=SECTION_HEADER_BG_COLOR,
                            end_color=SECTION_HEADER_BG_COLOR,
                            fill_type="solid",
                        )
                        ws.merge_cells(
                            start_row=row_num,
                            start_column=1,
                            end_row=row_num,
                            end_column=len(display_columns),
                        )
                        for merged_col in range(1, len(display_columns) + 1):
                            merged_cell = ws.cell(row=row_num, column=merged_col)
                            merged_cell.border = thin_border
                        break
                    else:
                        rich_text = parse_markdown_bold_to_rich_text(str(cell_value))
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.value = rich_text  # type: ignore
                        cell.fill = PatternFill(
                            start_color=CELL_BG_COLOR, end_color=CELL_BG_COLOR, fill_type="solid"
                        )
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        cell.border = thin_border

            # Insert images into their respective columns
            if first_image_column and image_id_to_path:
                _insert_images_into_excel(
                    ws,
                    study_table.rows,
                    image_id_to_path,
                    first_image_column,
                    start_row=2,
                    max_image_height=100,
                )
                # Set image column widths
                for i in range(max_images):
                    ws.column_dimensions[get_column_letter(first_image_column + i)].width = 20

            # Auto-adjust column widths (text columns only)
            for col_num in range(1, len(text_columns) + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            cell_length = len(str(cell.value)) if cell.value else 0
                            max_length = max(max_length, cell_length)
                        except Exception:
                            pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to temp file first, then upload
            with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name

            await asyncio.to_thread(wb.save, tmp_path)
            storage_path = await upload_job_file(tmp_path, user_id, job_id, output_filename)

            # Clean up temp file
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

        # Create artifact
        from core.models import ArtifactType

        await create_artifact(job_id, ArtifactType.EXCEL_STUDY_TABLE, storage_path)

        return {"xlsx_path": storage_path}
    except Exception as e:
        logging.exception(f"Failed to generate spreadsheet: {e}")
        return {}


@broker.task
async def generate_vignette_artifact_task(data: dict) -> dict:
    """Generate vignette PDF artifact as a distributed task.

    Returns dict with vignette_path if successful, empty dict if skipped/failed.
    """
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    user_id = ctx.user_id

    if not ctx.enable_vignette:
        return {}

    if ctx.outputs is None:
        return {}

    pdf_path = ctx.outputs.get("pdf_path")
    if not pdf_path:
        return {}

    try:
        # Download PDF if on S3 for AI processing
        async with temp_download(pdf_path) as local_pdf:
            vignette_data = await generate_vignette_questions(
                local_pdf,
                custom_prompt=ctx.vignette_prompt,
                user_id=user_id,
                job_id=job_id,
            )

        if not vignette_data.learning_objectives:
            return {}

        learning_objectives = [lo.model_dump() for lo in vignette_data.learning_objectives]

        template_path = settings.BASE_DIR / "templates" / "pdf"
        env = Environment(loader=FileSystemLoader(template_path), autoescape=select_autoescape())
        template = env.get_template("vignette.html")
        html = template.render(learning_objectives=learning_objectives)

        # Get base name from the PDF filename
        pdf_filename = ctx.outputs.get("pdf_filename", os.path.basename(pdf_path))
        base_name = os.path.splitext(pdf_filename)[0]
        output_filename = f"{base_name} - Vignette Questions.pdf"

        # Create PDF in temp file
        with NamedTemporaryFile(suffix=".pdf") as tmp:
            pisa_status = pisa.CreatePDF(html, dest=tmp)
            if hasattr(pisa_status, "err") and getattr(pisa_status, "err", None):
                return {}
            tmp.flush()
            storage_path = await upload_job_file(tmp.name, user_id, job_id, output_filename)

        # Create artifact
        from core.models import ArtifactType

        await create_artifact(job_id, ArtifactType.PDF_VIGNETTE, storage_path)

        return {"vignette_path": storage_path}
    except Exception as e:
        logging.exception(f"Failed to generate vignette: {e}")
        return {}


@broker.task
async def generate_mindmap_artifact_task(data: dict) -> dict:
    """Generate mindmap Mermaid file(s) as a distributed task.

    Returns dict with mindmap_paths (list) if successful, empty dict if skipped/failed.
    """
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    user_id = ctx.user_id

    if not ctx.enable_mindmap:
        return {}

    if ctx.outputs is None:
        return {}

    pdf_path = ctx.outputs.get("pdf_path")
    if not pdf_path:
        return {}

    try:
        # Download PDF if on S3 for AI processing
        async with temp_download(pdf_path) as local_pdf:
            mindmaps = await generate_mindmap(
                local_pdf,
                custom_prompt=ctx.mindmap_prompt,
                user_id=user_id,
                job_id=job_id,
            )

        if not mindmaps:
            return {}

        # Get base name from the PDF filename
        pdf_filename = ctx.outputs.get("pdf_filename", os.path.basename(pdf_path))
        base_name = os.path.splitext(pdf_filename)[0]
        mindmap_paths = []

        # Create artifact import
        from core.models import ArtifactType

        for i, (title, mermaid_code) in enumerate(mindmaps):
            # Sanitize title for filename
            safe_title = "".join(c for c in title if c.isalnum() or c in " -_").strip()
            if not safe_title:
                safe_title = f"Mindmap {i + 1}"

            mindmap_filename = f"{base_name} - {safe_title}.mmd"

            # Write to temp file and upload
            with NamedTemporaryFile(suffix=".mmd", mode="w", encoding="utf-8") as tmp:
                tmp.write(mermaid_code)
                tmp.flush()
                storage_path = await upload_job_file(tmp.name, user_id, job_id, mindmap_filename)

            # Create artifact for each mindmap
            await create_artifact(job_id, ArtifactType.MERMAID_MINDMAP, storage_path)
            mindmap_paths.append(storage_path)

        return {"mindmap_paths": mindmap_paths}
    except Exception as e:
        logging.exception(f"Failed to generate mindmap: {e}")
        return {}


@broker.task
async def generate_artifacts_task(data: dict) -> dict:
    """Generate all artifacts (spreadsheet, vignette, mindmap) in parallel using distributed tasks."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "generate_artifacts"

    await update_job_progress(job_id, stage_name, 0, "Generating artifacts")

    if ctx.outputs is None:
        ctx.outputs = {}

    pdf_path = ctx.outputs.get("pdf_path")
    if not pdf_path:
        return ctx.to_dict()

    # Queue all tasks with their names for progress reporting
    tasks_with_names = [
        (await generate_spreadsheet_artifact_task.kiq(data), "Spreadsheet"),
        (await generate_vignette_artifact_task.kiq(data), "Vignette"),
        (await generate_mindmap_artifact_task.kiq(data), "Mindmap"),
    ]

    await update_job_progress(job_id, stage_name, 0.1, "Waiting for artifact tasks")

    # Wait for results as they complete and report progress
    completed_count = 0
    total_tasks = len(tasks_with_names)

    # Create futures for each task's wait_result
    async def wait_and_identify(task_handle, name: str):
        result = await task_handle.wait_result()
        return (result, name)

    pending = [
        asyncio.create_task(wait_and_identify(task, name)) for task, name in tasks_with_names
    ]

    while pending:
        done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)

        for completed_task in done:
            try:
                result, task_name = await completed_task
                completed_count += 1

                # Calculate progress (0.1 to 0.9 range for artifact generation)
                progress = 0.1 + (completed_count / total_tasks) * 0.8

                # Check if task succeeded
                if result is not None and not result.is_err:
                    return_value = getattr(result, "return_value", None)
                    if isinstance(return_value, dict) and return_value:
                        ctx.outputs.update(return_value)
                        await update_job_progress(
                            job_id,
                            stage_name,
                            progress,
                            f"{task_name} completed ({completed_count}/{total_tasks})",
                        )
                    else:
                        await update_job_progress(
                            job_id,
                            stage_name,
                            progress,
                            f"{task_name} skipped ({completed_count}/{total_tasks})",
                        )
                else:
                    await update_job_progress(
                        job_id,
                        stage_name,
                        progress,
                        f"{task_name} failed ({completed_count}/{total_tasks})",
                    )
                    if result is not None and result.is_err:
                        logging.error(f"{task_name} artifact generation failed: {result.error}")

            except Exception as e:
                completed_count += 1
                progress = 0.1 + (completed_count / total_tasks) * 0.8
                logging.exception(f"Artifact task failed: {e}")
                await update_job_progress(
                    job_id, stage_name, progress, f"Task failed ({completed_count}/{total_tasks})"
                )

        # Convert remaining pending set back to list for next iteration
        pending = set(pending)

    await update_job_progress(job_id, stage_name, 1.0, "Artifacts generated")

    return ctx.to_dict()
