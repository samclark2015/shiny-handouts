"""
Taskiq tasks for the video processing pipeline.

Converts the pipeline stages into Taskiq tasks with:
- Progress reporting via Redis pub/sub
- Stage-level caching via Redis
- Pipeline chaining via taskiq-pipelines
"""

import asyncio
import hashlib
import json
import logging
import os
import shutil
import subprocess
import tempfile
import urllib.request
from contextlib import contextmanager
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from tempfile import TemporaryDirectory
from typing import cast
from urllib.parse import urljoin
from uuid import uuid4

import cv2
import m3u8
import pandas as pd
import redis.asyncio as aioredis
import skimage as ski
from asgiref.sync import sync_to_async
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from taskiq import TaskiqEvents
from taskiq_pipelines import Pipeline, PipelineMiddleware
from taskiq_redis import ListQueueBroker, RedisAsyncResultBackend
from xhtml2pdf import pisa

# Import AI functions - they will run in async context
from pipeline.ai import (
    clean_transcript,
    generate_captions,
    generate_spreadsheet_helper,
    generate_title,
    generate_vignette_questions,
)
from pipeline.helpers import Caption, Slide, fetch, parse_markdown_bold_to_rich_text

from .cache import CacheContext
from .middleware import PipelineErrorMiddleware

# Directory configuration
IN_DIR = os.path.join("data", "input")
OUT_DIR = os.path.join("data", "output")
FRAMES_DIR = os.path.join("data", "frames")

# Redis configuration
REDIS_URL = os.environ.get("REDIS_URL", "redis://localhost:6379/0")

# Create the result backend
result_backend = RedisAsyncResultBackend(
    redis_url=REDIS_URL,
    result_ex_time=60 * 60 * 24,  # Results expire after 24 hours
)

# Create the broker with middlewares
broker = (
    ListQueueBroker(
        url=REDIS_URL,
        queue_name="handout_generator",
    )
    .with_result_backend(result_backend)
    .with_middlewares(
        PipelineErrorMiddleware(),
        PipelineMiddleware(),
    )
)


# Ensure directories exist
os.makedirs(IN_DIR, exist_ok=True)
os.makedirs(OUT_DIR, exist_ok=True)
os.makedirs(FRAMES_DIR, exist_ok=True)


# Pipeline stage weights (must sum to 1.0)
STAGE_WEIGHTS = {
    "generate_context": 0.02,  # 2%
    "download_video": 0.15,  # 15%
    "extract_captions": 0.15,  # 15%
    "match_frames": 0.15,  # 15%
    "transform_slides_with_ai": 0.15,  # 15%
    "generate_output": 0.10,  # 10%
    "compress_pdf": 0.08,  # 8%
    "generate_spreadsheet": 0.10,  # 10%
    "generate_vignette_pdf": 0.08,  # 8%
    "finalize_job": 0.02,  # 2%
}

# Calculate cumulative stage start positions
STAGE_START_PROGRESS = {}
_cumulative = 0.0
for stage, weight in STAGE_WEIGHTS.items():
    STAGE_START_PROGRESS[stage] = _cumulative
    _cumulative += weight


def calculate_overall_progress(stage_name: str, stage_progress: float) -> float:
    """Calculate overall pipeline progress from stage name and stage-specific progress.

    Args:
        stage_name: Name of the current stage (e.g., 'download_video')
        stage_progress: Progress within this stage (0.0 to 1.0)

    Returns:
        Overall progress from 0.0 to 1.0
    """
    if stage_name not in STAGE_WEIGHTS:
        return 0.0

    start = STAGE_START_PROGRESS[stage_name]
    weight = STAGE_WEIGHTS[stage_name]
    return start + (stage_progress * weight)


class JobCancelledException(Exception):
    """Raised when a job has been cancelled."""

    pass


async def check_job_cancelled(job_id: int) -> bool:
    """Check if a job has been cancelled or is being cancelled."""
    from core.models import Job, JobStatus

    try:
        job = await Job.objects.aget(id=job_id)
        return job.status in (JobStatus.CANCELLING, JobStatus.CANCELLED)
    except Job.DoesNotExist:
        return True  # Treat missing job as cancelled


@dataclass
class TaskContext:
    """Context object that flows through the pipeline stages."""

    job_id: int
    source_id: str
    input_type: str  # 'url', 'upload', 'panopto'
    input_data: dict  # Serialized input configuration
    use_ai: bool = True
    video_path: str | None = None
    captions: list[dict] | None = None
    slides: list[dict] | None = None
    outputs: dict | None = None

    # Per-job settings
    enable_excel: bool = True
    enable_vignette: bool = True

    # User settings (custom prompts)
    vignette_prompt: str | None = None
    spreadsheet_prompt: str | None = None
    spreadsheet_columns: list[dict] | None = None

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> "TaskContext":
        return cls(**data)

    def update_from(self, other: "TaskContext") -> None:
        """Update this instance with non-None values from another instance."""
        for field in self.__dataclass_fields__:
            other_value = getattr(other, field)
            if other_value is not None:
                setattr(self, field, other_value)

    def get_video_path(self) -> str:
        """Get the path to the video file."""
        if self.video_path and os.path.exists(self.video_path):
            return self.video_path
        return os.path.join(IN_DIR, f"video_{self.source_id}.mp4")


async def publish_progress(job_id: int, stage: str, progress: float, message: str = "") -> None:
    """Publish progress update to Redis pub/sub."""
    redis = await aioredis.from_url(REDIS_URL)
    try:
        await redis.publish(
            f"job:{job_id}:progress",
            json.dumps(
                {
                    "stage": stage,
                    "progress": progress,
                    "message": message,
                    "status": "running",
                }
            ),
        )
    finally:
        await redis.close()


async def update_job_progress(job_id: int, stage: str, progress: float, message: str) -> None:
    """Update job progress in the database and publish to Redis.

    Also checks for cancellation and raises JobCancelledException if cancelled.

    Args:
        job_id: The job ID to update
        stage: The current stage name (e.g., 'download_video')
        progress: Progress within the current stage (0.0 to 1.0)
        message: Human-readable progress message
    """
    from core.models import Job, JobStatus

    # Calculate overall progress across all stages
    overall_progress = calculate_overall_progress(stage, progress)

    try:
        job = await Job.objects.aget(id=job_id)

        # Check for cancellation
        if job.status == JobStatus.CANCELLING:
            job.status = JobStatus.CANCELLED
            job.completed_at = datetime.now(UTC)
            await job.asave(update_fields=["status", "completed_at"])

            raise JobCancelledException(f"Job {job_id} was cancelled")

        job.current_stage = message
        job.progress = overall_progress
        if job.status == JobStatus.PENDING:
            job.status = JobStatus.RUNNING
            job.started_at = datetime.now(UTC)
        await job.asave(update_fields=["current_stage", "progress", "status", "started_at"])
    except Job.DoesNotExist:
        raise JobCancelledException(f"Job {job_id} no longer exists") from None

    await publish_progress(job_id, stage, overall_progress, message)


async def update_job_label(job_id: int, label: str) -> None:
    """Update job label in the database."""
    from core.models import Job

    try:
        job = await Job.objects.aget(id=job_id)
        job.label = label
        await job.asave(update_fields=["label"])
    except Job.DoesNotExist:
        pass


@sync_to_async
def get_or_create_lecture(job_id: int, source_id: str | None = None):
    """Get or create lecture for a job, keyed by source_id.

    Args:
        job_id: The job ID
        source_id: Optional source ID. If not provided, will try to get from job's lecture.
    """
    from core.models import Job, Lecture

    job = Job.objects.select_related("user").get(id=job_id)

    # If source_id not provided, try to get from existing lecture
    if not source_id:
        try:
            existing_lecture = Lecture.objects.get(job=job)
            source_id = existing_lecture.source_id
        except Lecture.DoesNotExist:
            # No source_id and no existing lecture, create with empty source_id
            source_id = ""

    # Check if lecture already exists for this source_id and user
    if source_id:
        try:
            lecture = Lecture.objects.get(source_id=source_id, user=job.user)
            # Update the job reference if this is a retry
            if lecture.job_id != job.pk:
                lecture.job_id = job.pk
                lecture.save(update_fields=["job"])
            return lecture
        except Lecture.DoesNotExist:
            pass

    # Check if lecture exists by job (for backwards compatibility)
    try:
        lecture = Lecture.objects.get(job=job)
        # Update source_id if we have one now
        if source_id and not lecture.source_id:
            lecture.source_id = source_id
            lecture.save(update_fields=["source_id"])
        return lecture
    except Lecture.DoesNotExist:
        pass

    # Create new lecture
    lecture = Lecture(
        user=job.user,
        job=job,
        title=job.label,
        source_id=source_id,
        date=datetime.now(UTC),
    )
    lecture.save()
    return lecture


@sync_to_async
def create_artifact(
    job_id: int, artifact_type, file_path: str, source_id: str | None = None
) -> None:
    """Create an artifact record immediately when a file is generated.

    Args:
        job_id: The job ID
        artifact_type: Type of artifact (PDF_HANDOUT, EXCEL_STUDY_TABLE, PDF_VIGNETTE)
        file_path: Path to the generated file
        source_id: Optional source ID for lecture lookup
    """
    from core.models import Artifact

    if not file_path or not os.path.exists(file_path):
        return

    try:
        # Call the sync version directly since we're in sync context
        from core.models import Job, Lecture

        job = Job.objects.select_related("user").get(id=job_id)

        # If source_id not provided, try to get from existing lecture
        if not source_id:
            try:
                existing_lecture = Lecture.objects.get(job=job)
                source_id = existing_lecture.source_id
            except Lecture.DoesNotExist:
                source_id = ""

        # Check if lecture already exists for this source_id and user
        lecture = None
        if source_id:
            try:
                lecture = Lecture.objects.get(source_id=source_id, user=job.user)
                if lecture.job != job.pk:
                    lecture.job = job
                    lecture.save(update_fields=["job"])
            except Lecture.DoesNotExist:
                pass

        # Check if lecture exists by job (for backwards compatibility)
        if not lecture:
            try:
                lecture = Lecture.objects.get(job=job)
                if source_id and not lecture.source_id:
                    lecture.source_id = source_id
                    lecture.save(update_fields=["source_id"])
            except Lecture.DoesNotExist:
                pass

        # Create new lecture if needed
        if not lecture:
            lecture = Lecture.objects.create(
                user=job.user,
                job=job,
                title=job.label,
                source_id=source_id,
                date=datetime.now(UTC),
            )

        # Check if artifact already exists for this lecture and type
        existing = Artifact.objects.filter(lecture=lecture, artifact_type=artifact_type).first()

        if existing:
            # Update existing artifact
            existing.file_path = file_path
            existing.file_name = os.path.basename(file_path)
            existing.file_size = os.path.getsize(file_path)
            existing.save()
        else:
            # Create new artifact
            Artifact.objects.create(
                lecture=lecture,
                artifact_type=artifact_type,
                file_path=file_path,
                file_name=os.path.basename(file_path),
                file_size=int(os.path.getsize(file_path)),
            )
    except Exception as e:
        # Log error but don't fail the task
        logging.exception(f"Failed to create artifact for job {job_id}: {e}")


async def mark_job_completed(job_id: int, outputs: dict) -> None:
    """Mark a job as completed in the database."""
    from core.models import Job, JobStatus, Lecture

    try:
        job = await Job.objects.select_related("user").aget(id=job_id)
        job.status = JobStatus.COMPLETED
        job.progress = 1.0
        job.completed_at = datetime.now(UTC)
        await job.asave(update_fields=["status", "progress", "completed_at"])

        # Update lecture source_id if it exists
        try:
            lecture = await Lecture.objects.aget(job=job)
            if "source_id" in outputs:
                lecture.source_id = outputs["source_id"]
                await lecture.asave(update_fields=["source_id"])
        except Lecture.DoesNotExist:
            pass

        # Publish completion
        await publish_progress(job_id, "completed", 1.0, "Job completed successfully")

    except Job.DoesNotExist:
        pass


@contextmanager
def cache_context(ctx: TaskContext, stage_name: str):
    """Context manager for stage-level caching.

    Automatically updates ctx from cache if available and yields True.
    Otherwise yields False and stores ctx in cache on exit.

    Usage:
        with cache_context(ctx, stage_name) as cache_hit:
            if cache_hit:
                return  # Skip execution, data loaded from cache
            # ... do work ...
    """
    cache = CacheContext(ctx.source_id)

    # Check cache on entry
    cached = cache.get(stage_name)

    if cached:
        logging.info(f"Cache hit for stage {stage_name}, loading cached data")
        ctx.update_from(TaskContext.from_dict({**cached, "job_id": ctx.job_id}))

    yield bool(cached)  # Yield True if cache hit, False otherwise

    # Store in cache on exit only if it wasn't cached
    if not cached:
        cache.set(stage_name, ctx.to_dict())


# Worker lifecycle events


@broker.on_event(TaskiqEvents.WORKER_STARTUP)
async def on_worker_startup(state):
    """Initialize resources when worker starts."""
    # Initialize Django before using any models
    import django

    django.setup()

    # Store Redis connection for pub/sub in state
    state.redis = await aioredis.from_url(REDIS_URL)


@broker.on_event(TaskiqEvents.WORKER_SHUTDOWN)
async def on_worker_shutdown(state):
    """Cleanup when worker shuts down."""
    if hasattr(state, "redis"):
        await state.redis.close()


# Pipeline stage tasks


@broker.task
async def generate_context_task(job_id: int, input_type: str, input_data: str) -> dict:
    """Generate processing context from input."""
    from core.models import Job

    stage_name = "generate_context"

    await update_job_progress(job_id, stage_name, 0, "Initializing")

    input_dict = cast(dict, json.loads(input_data))
    # Generate source_id based on input type
    if input_type == "panopto":
        source_id = input_dict.get("delivery_id", "")
    elif input_type == "url":
        source_id = input_dict.get("url", "")
    else:  # upload
        source_id = await _hash_file(input_dict.get("path", ""))

    # Load job settings and profile
    job = await Job.objects.select_related("user", "setting_profile").aget(id=job_id)
    enable_excel = job.enable_excel
    enable_vignette = job.enable_vignette

    # Load settings from profile (if set)
    vignette_prompt = None
    spreadsheet_prompt = None
    spreadsheet_columns = None

    if job.setting_profile:
        vignette_prompt = job.setting_profile.get_vignette_prompt()
        spreadsheet_prompt = job.setting_profile.get_spreadsheet_prompt()
        spreadsheet_columns = job.setting_profile.get_spreadsheet_columns()

    ctx = TaskContext(
        job_id=job_id,
        source_id=source_id,
        input_type=input_type,
        input_data=input_dict,
        use_ai=True,
        enable_excel=enable_excel,
        enable_vignette=enable_vignette,
        vignette_prompt=vignette_prompt,
        spreadsheet_prompt=spreadsheet_prompt,
        spreadsheet_columns=spreadsheet_columns,
    )

    await update_job_progress(job_id, stage_name, 1.0, "Context created")

    return ctx.to_dict()


@broker.task
async def download_video_task(data: dict) -> dict:
    """Download video if it doesn't exist."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "download_video"

    await update_job_progress(job_id, stage_name, 0, "Checking video")

    # Check cache first
    with cache_context(ctx, stage_name) as cache_hit:
        if cache_hit:
            await update_job_progress(job_id, stage_name, 1.0, "Video downloaded")
            return ctx.to_dict()

        video_path = ctx.get_video_path()

        if ctx.input_type == "upload":
            upload_path = ctx.input_data.get("path", "")
            if os.path.exists(upload_path):
                ctx.video_path = upload_path
                await update_job_progress(job_id, stage_name, 1.0, "Video ready")
                return ctx.to_dict()

        if ctx.input_type == "panopto":
            await _download_panopto_video(job_id, stage_name, ctx.input_data, video_path)
        else:
            video_url = ctx.input_data.get("url", "")
            if _is_m3u8_url(video_url):
                await _download_m3u8_stream(job_id, stage_name, video_url, video_path)
            else:
                await _download_regular_video(job_id, stage_name, video_url, video_path)

            # Hash the downloaded file contents to generate source_id
            ctx.source_id = await _hash_file(video_path)

        ctx.video_path = video_path

    await update_job_progress(job_id, stage_name, 1.0, "Video downloaded")

    return ctx.to_dict()


@broker.task
async def extract_captions_task(data: dict) -> dict:
    """Extract captions from the video."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "extract_captions"

    await update_job_progress(job_id, stage_name, 0, "Extracting captions")

    with cache_context(ctx, stage_name) as cache_hit:
        if cache_hit:
            await update_job_progress(job_id, stage_name, 1.0, "Captions extracted")
            return ctx.to_dict()

        captions = await generate_captions(ctx.get_video_path())
        ctx.captions = [{"text": c.text, "timestamp": c.timestamp} for c in captions]

    await update_job_progress(job_id, stage_name, 1.0, "Captions extracted")

    return ctx.to_dict()


@broker.task
async def match_frames_task(data: dict) -> dict:
    """Match frames to captions based on structural similarity."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "match_frames"

    await update_job_progress(job_id, stage_name, 0, "Matching frames")

    # Check cache first
    with cache_context(ctx, stage_name) as cache_hit:
        if cache_hit:
            await update_job_progress(job_id, stage_name, 1.0, "Frames matched")
            return ctx.to_dict()

        if not ctx.captions:
            ctx.slides = []
            return ctx.to_dict()

        captions = [Caption(**c) for c in ctx.captions]
        last_frame = None
        last_frame_gs = None
        cum_captions = []
        pairs = []

        stream = cv2.VideoCapture()
        stream.open(ctx.get_video_path())

        frame_path = os.path.join(FRAMES_DIR, ctx.source_id)
        os.makedirs(frame_path, exist_ok=True)

        for idx, cap in enumerate(captions):
            stream.set(cv2.CAP_PROP_POS_MSEC, cap.timestamp * 1_000 + 500)
            ret, frame = stream.read()
            if not ret:
                continue

            frame_gs = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            if last_frame is None:
                last_frame = frame
                last_frame_gs = frame_gs
                cum_captions.append(cap.text)
                continue

            similarity_result = ski.metrics.structural_similarity(
                last_frame_gs, frame_gs, full=False
            )
            score = (
                similarity_result
                if isinstance(similarity_result, (int, float))
                else similarity_result[0]
            )

            if score < 0.925 or (idx + 1) == len(captions):
                cap_full = " ".join(cum_captions)
                image_path = os.path.join(frame_path, f"{uuid4()}.png")
                cv2.imwrite(image_path, last_frame)

                pairs.append({"image": image_path, "caption": cap_full, "extra": None})
                last_frame = frame
                last_frame_gs = frame_gs
                cum_captions.clear()

                progress = (idx + 1) / len(captions)
                await update_job_progress(job_id, stage_name, progress * 0.9, "Matching slides")

            cum_captions.append(cap.text)

        stream.release()
        ctx.slides = pairs

    await update_job_progress(job_id, stage_name, 1.0, "Frames matched")

    return ctx.to_dict()


@broker.task
async def transform_slides_ai_task(data: dict) -> dict:
    """Apply AI transformation to slides."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "transform_slides_with_ai"

    await update_job_progress(job_id, stage_name, 0, "Transforming slides with AI")

    if not ctx.use_ai or not ctx.slides:
        return ctx.to_dict()

    # Check cache first
    with cache_context(ctx, stage_name) as cache_hit:
        if cache_hit:
            await update_job_progress(job_id, stage_name, 1.0, "Slides transformed")
            return ctx.to_dict()

        output = []
        slides = cast(list[dict], ctx.slides)
        total = len(slides)

        for idx, slide in enumerate(slides):
            cleaned = await clean_transcript(slide["caption"])
            output.append(
                {
                    "image": slide["image"],
                    "caption": cleaned,
                    "extra": slide.get("extra"),
                }
            )

            progress = (idx + 1) / total
            await update_job_progress(job_id, stage_name, progress * 0.9, "Cleaning transcript")

        ctx.slides = output

    await update_job_progress(job_id, stage_name, 1.0, "Slides transformed")

    return ctx.to_dict()


@broker.task
async def generate_output_task(data: dict) -> dict:
    """Generate the PDF output."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "generate_output"

    await update_job_progress(job_id, stage_name, 0, "Generating PDF")

    # Check cache first
    with cache_context(ctx, stage_name) as cache_hit:
        if cache_hit:
            await update_job_progress(job_id, stage_name, 1.0, "PDF generated")
            return ctx.to_dict()

        # Convert slide dicts back to Slide namedtuples for template
        slides = [Slide(**s) for s in ctx.slides] if ctx.slides else []

        template_path = os.path.join(os.path.dirname(__file__), "..", "templates", "pdf")
        env = Environment(loader=FileSystemLoader(template_path), autoescape=select_autoescape())
        template = env.get_template("template.html")
        html = template.render(pairs=slides)

        await update_job_progress(job_id, stage_name, 0.3, "Generating title")
        title = await generate_title(html)
        await update_job_label(job_id, title)

        path = os.path.join(OUT_DIR, f"{title}.pdf")
        os.makedirs(OUT_DIR, exist_ok=True)

        await update_job_progress(job_id, stage_name, 0.5, "Creating PDF")
        with open(path, "wb") as f:
            pisa_status = pisa.CreatePDF(html, dest=f)
            if hasattr(pisa_status, "err") and getattr(pisa_status, "err", None):
                raise ValueError("Error generating PDF")

        ctx.outputs = ctx.outputs or {}
        ctx.outputs["pdf_path"] = path
        ctx.outputs["source_id"] = ctx.source_id

    await update_job_progress(job_id, stage_name, 1.0, "PDF generated")

    return ctx.to_dict()


@broker.task
async def compress_pdf_task(data: dict) -> dict:
    """Compress the PDF using Ghostscript."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "compress_pdf"

    await update_job_progress(job_id, stage_name, 0, "Compressing PDF")

    pdf_path = ctx.outputs.get("pdf_path") if ctx.outputs else None
    if not pdf_path or not os.path.exists(pdf_path):
        return ctx.to_dict()

    source_id = data.get("source_id", "")

    with cache_context(ctx, stage_name) as cache_hit:
        if cache_hit:
            await update_job_progress(job_id, stage_name, 1.0, "PDF compressed")
            return ctx.to_dict()

        # Compress using Ghostscript
        with TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, f"compressed_{os.path.basename(pdf_path)}")

            gs_command = [
                "gs",
                "-sDEVICE=pdfwrite",
                "-dCompatibilityLevel=1.4",
                "-dPDFSETTINGS=/ebook",
                "-dNOPAUSE",
                "-dQUIET",
                "-dBATCH",
                f"-sOutputFile={output_path}",
                pdf_path,
            ]

            try:
                await asyncio.to_thread(subprocess.run, gs_command, check=True)
                shutil.move(output_path, pdf_path)
            except (subprocess.CalledProcessError, FileNotFoundError) as e:
                print(f"Ghostscript compression failed: {e}")

        # Create artifact immediately
        from core.models import ArtifactType

        await create_artifact(job_id, ArtifactType.PDF_HANDOUT, pdf_path, source_id)

        await update_job_progress(job_id, stage_name, 1.0, "PDF compressed")

    return ctx.to_dict()


@broker.task
async def generate_spreadsheet_task(data: dict) -> dict:
    """Generate the Excel spreadsheet."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "generate_spreadsheet"

    # Skip if Excel generation is disabled
    if not ctx.enable_excel:
        await update_job_progress(job_id, stage_name, 1.0, "Excel generation skipped")
        return ctx.to_dict()

    await update_job_progress(job_id, stage_name, 0, "Generating spreadsheet")

    if ctx.outputs is None:
        ctx.outputs = {}

    pdf_path = ctx.outputs.get("pdf_path")
    if not pdf_path or not os.path.exists(pdf_path):
        return ctx.to_dict()

    with cache_context(ctx, stage_name) as cache_hit:
        if cache_hit:
            await update_job_progress(job_id, stage_name, 1.0, "Spreadsheet generated")
            return ctx.to_dict()

        study_table = await generate_spreadsheet_helper(
            pdf_path,
            custom_prompt=ctx.spreadsheet_prompt,
            custom_columns=ctx.spreadsheet_columns,
        )

        if not study_table.rows:
            raise ValueError("No rows found in data")

        df = pd.DataFrame(study_table.rows)

        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        output_filename = os.path.join(OUT_DIR, f"{base_name}.xlsx")

        await update_job_progress(job_id, stage_name, 0.6, "Writing Excel file")

        # Style constants - modify these to change colors
        HEADER_BG_COLOR = "D3D3D3"  # Light grey
        CELL_BG_COLOR = "ADD8E6"  # Baby blue
        SECTION_HEADER_BG_COLOR = "6CB4E8"  # Darker blue for single-cell rows
        BORDER_COLOR = "000000"  # Black

        # Create border style
        thin_border = Border(
            left=Side(style="thin", color=BORDER_COLOR),
            right=Side(style="thin", color=BORDER_COLOR),
            top=Side(style="thin", color=BORDER_COLOR),
            bottom=Side(style="thin", color=BORDER_COLOR),
        )

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws.title = "Study Table"

        # Write header row
        for col_num, column_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill(
                start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid"
            )
            cell.border = thin_border

        # Write data rows
        for row_num, row_data in enumerate(study_table.rows, 2):
            # Check if this is a single-cell row (only first cell has content)
            non_empty_cells = sum(1 for col_name in df.columns if row_data.get(col_name, ""))
            is_section_header = non_empty_cells == 1

            for col_num, column_name in enumerate(df.columns, 1):
                cell_value = row_data.get(column_name, "")
                rich_text_value = parse_markdown_bold_to_rich_text(cell_value)
                cell = ws.cell(row=row_num, column=col_num)
                # Only set value if cell is not a merged cell
                cell.value = rich_text_value  # pyright: ignore[reportAttributeAccessIssue]
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

                # Apply background color and font styling
                if is_section_header:
                    cell.fill = PatternFill(
                        start_color=SECTION_HEADER_BG_COLOR,
                        end_color=SECTION_HEADER_BG_COLOR,
                        fill_type="solid",
                    )
                    if cell_value:  # Only make bold if there's content
                        cell.font = Font(bold=True, size=11)
                else:
                    cell.fill = PatternFill(
                        start_color=CELL_BG_COLOR, end_color=CELL_BG_COLOR, fill_type="solid"
                    )

        # Auto-adjust column widths
        for col_num in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        cell_length = min(len(str(cell.value)), 100)
                        max_length = max(max_length, cell_length)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

        await asyncio.to_thread(wb.save, output_filename)

        ctx.outputs["xlsx_path"] = output_filename

        # Create artifact immediately
        from core.models import ArtifactType

        await create_artifact(
            job_id, ArtifactType.EXCEL_STUDY_TABLE, output_filename, ctx.source_id
        )

        await update_job_progress(job_id, stage_name, 1.0, "Spreadsheet generated")

    return ctx.to_dict()


@broker.task
async def generate_vignette_task(data: dict) -> dict:
    """Generate the vignette PDF."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id
    stage_name = "generate_vignette_pdf"

    # Skip if vignette generation is disabled
    if not ctx.enable_vignette:
        await update_job_progress(job_id, stage_name, 1.0, "Vignette generation skipped")
        return ctx.to_dict()

    await update_job_progress(job_id, stage_name, 0, "Generating vignette questions")

    if ctx.outputs is None:
        ctx.outputs = {}

    pdf_path = ctx.outputs.get("pdf_path")
    if not pdf_path or not os.path.exists(pdf_path):
        return ctx.to_dict()

    await update_job_progress(job_id, stage_name, 0.2, "Generating questions")
    vignette_data = await generate_vignette_questions(
        pdf_path,
        custom_prompt=ctx.vignette_prompt,
    )

    if not vignette_data.learning_objectives:
        raise ValueError("No learning objectives found")

    learning_objectives = [lo.model_dump() for lo in vignette_data.learning_objectives]

    template_path = os.path.join(os.path.dirname(__file__), "..", "templates", "pdf")
    env = Environment(loader=FileSystemLoader(template_path), autoescape=select_autoescape())
    template = env.get_template("vignette.html")
    html = template.render(learning_objectives=learning_objectives)

    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    vignette_pdf_path = os.path.join(OUT_DIR, f"{base_name} - Vignette Questions.pdf")

    await update_job_progress(job_id, stage_name, 0.7, "Creating vignette PDF")

    with open(vignette_pdf_path, "wb") as f:
        pisa_status = pisa.CreatePDF(html, dest=f)
        if hasattr(pisa_status, "err") and getattr(pisa_status, "err", None):
            raise ValueError("Error generating vignette PDF")

    ctx.outputs["vignette_path"] = vignette_pdf_path

    # Create artifact immediately
    from core.models import ArtifactType

    await create_artifact(job_id, ArtifactType.PDF_VIGNETTE, vignette_pdf_path, ctx.source_id)

    await update_job_progress(job_id, stage_name, 1.0, "Vignette generated")

    return ctx.to_dict()


@broker.task
async def finalize_job_task(data: dict) -> dict:
    """Finalize the job and create database records."""
    ctx = TaskContext.from_dict(data)
    job_id = ctx.job_id

    outputs = ctx.outputs or {}
    outputs["source_id"] = ctx.source_id

    await mark_job_completed(job_id, outputs)

    return {"job_id": job_id, "status": "completed", "outputs": outputs}


# Helper functions for video download


async def _hash_file(file_path: str) -> str:
    """Generate SHA256 hash of file contents.

    Args:
        file_path: Path to the file to hash

    Returns:
        Hexadecimal hash string
    """

    def _hash_sync():
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            # Read file in chunks to handle large files
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()

    return await asyncio.to_thread(_hash_sync)


def _is_m3u8_url(url: str) -> bool:
    """Check if a URL points to an M3U8 file."""
    return url.endswith(".m3u8") or "m3u8" in url


async def _download_regular_video(
    job_id: int, stage_name: str, video_url: str, video_path: str
) -> None:
    """Download a regular video file."""

    def download():
        def report_progress(count, block_size, total_size):
            # Note: Can't easily update from sync callback in threaded context
            pass

        opener = urllib.request.build_opener()
        opener.addheaders = [("Range", "bytes=0-")]
        urllib.request.install_opener(opener)

        urllib.request.urlretrieve(video_url, video_path, reporthook=report_progress)

    await update_job_progress(job_id, stage_name, 0.1, "Downloading video")
    await asyncio.to_thread(download)
    await update_job_progress(job_id, stage_name, 0.9, "Download complete")


async def _download_m3u8_stream(
    job_id: int, stage_name: str, video_url: str, video_path: str
) -> None:
    """Download and combine M3U8 stream segments."""
    await update_job_progress(job_id, stage_name, 0.1, "Parsing playlist")

    playlist = await asyncio.to_thread(m3u8.load, video_url)

    if playlist.is_variant:
        if playlist.playlists:
            best_playlist = min(
                playlist.playlists,
                key=lambda p: p.stream_info.bandwidth if p.stream_info.bandwidth else 0,
            )
            stream_url = urljoin(video_url, best_playlist.uri)
            playlist = await asyncio.to_thread(m3u8.load, stream_url)
        else:
            raise ValueError("No streams found in variant playlist")

    segments = playlist.segments
    total_segments = len(segments)

    if total_segments == 0:
        raise ValueError("No segments found in playlist")

    with tempfile.TemporaryDirectory() as temp_dir:
        segment_files = []

        for i, segment in enumerate(segments):
            segment_url = urljoin(playlist.base_uri or video_url, segment.uri)
            segment_path = os.path.join(temp_dir, f"segment_{i:04d}.ts")

            for attempt in range(3):
                try:
                    await asyncio.to_thread(urllib.request.urlretrieve, segment_url, segment_path)
                    if os.path.getsize(segment_path) > 0:
                        break
                except Exception as e:
                    if attempt == 2:
                        raise ValueError(f"Failed to download segment {i}: {e}") from e

            segment_files.append(segment_path)
            progress = (i + 1) / total_segments * 0.8
            await update_job_progress(job_id, stage_name, progress, "Downloading segments")

        await update_job_progress(job_id, stage_name, 0.85, "Combining segments")

        concat_file = os.path.join(temp_dir, "segments.txt")
        with open(concat_file, "w") as f:
            for segment_file in segment_files:
                f.write(f"file '{segment_file}'\n")

        result = await asyncio.to_thread(
            subprocess.run,
            [
                "ffmpeg",
                "-f",
                "concat",
                "-safe",
                "0",
                "-i",
                concat_file,
                "-c",
                "copy",
                "-y",
                video_path,
            ],
            capture_output=True,
            text=True,
        )

        if result.returncode != 0:
            with open(video_path, "wb") as outfile:
                for segment_file in segment_files:
                    with open(segment_file, "rb") as infile:
                        outfile.write(infile.read())


async def _download_panopto_video(
    job_id: int, stage_name: str, panopto_data: dict, video_path: str
) -> None:
    """Download video from Panopto."""
    base = panopto_data["base"]
    cookie = panopto_data["cookie"]
    delivery_id = panopto_data["delivery_id"]

    await update_job_progress(job_id, stage_name, 0.1, "Getting Panopto info")

    delivery_info = await asyncio.to_thread(
        fetch,
        base,
        cookie,
        "Panopto/Pages/Viewer/DeliveryInfo.aspx",
        {
            "deliveryId": delivery_id,
            "responseType": "json",
            "getCaptions": "false",
            "language": "0",
        },
    )

    vidurl = delivery_info["Delivery"]["PodcastStreams"][0]["StreamUrl"]

    if _is_m3u8_url(vidurl):
        await _download_m3u8_stream(job_id, stage_name, vidurl, video_path)
    else:
        await _download_regular_video(job_id, stage_name, vidurl, video_path)


def create_pipeline(job_id: int, input_type: str, input_data: str) -> Pipeline:
    """Create a taskiq Pipeline for the full processing chain."""
    return (
        Pipeline(broker, generate_context_task)
        .call_next(download_video_task)
        .call_next(extract_captions_task)
        .call_next(match_frames_task)
        .call_next(transform_slides_ai_task)
        .call_next(generate_output_task)
        .call_next(compress_pdf_task)
        .call_next(generate_spreadsheet_task)
        .call_next(generate_vignette_task)
        .call_next(finalize_job_task)
    )


async def start_pipeline(job_id: int, input_type: str, input_data: str) -> str:
    """Start the pipeline and return the Taskiq task ID."""
    pipeline = create_pipeline(job_id, input_type, input_data)
    task = await pipeline.kiq(job_id, input_type, input_data)
    return task.task_id
